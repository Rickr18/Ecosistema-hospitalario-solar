"""
Actividad 2: Modelo básico de generación solar en Python
Hospital Nazareth – Barranquilla, Colombia

Calcula la generación solar fotovoltaica diaria estimada usando:
  - Irradiancia solar promedio diaria (kWh/m²/día)
  - Área disponible para paneles solares (m²)
  - Eficiencia del sistema fotovoltaico (fracción decimal)

Fórmula base:
  Energía generada (kWh/día) = Irradiancia × Área × Eficiencia

Incluye lectura del consumo real de Hospital Nazareth 1 (sede 1)
a partir del archivo Excel de consumo de energía 2018.
"""

import math
import os
import sys

import openpyxl
from playwright.sync_api import sync_playwright

# ---------------------------------------------------------------------------
# Parámetros del Hospital Nazareth (Barranquilla, Colombia)
# ---------------------------------------------------------------------------
SEDE_OBJETIVO = "HOSPITAL NAZARETH 1"   # Única sede analizada (instrucción docente)
AREA_PANELES_M2 = 500.0          # Área disponible para instalación de paneles (m²)
EFICIENCIA_SISTEMA = 0.15        # Eficiencia del sistema FV (15 %)
CONSUMO_DIARIO_KWH = 2800.0      # Consumo diario estimado del hospital (kWh/día)
FRACCION_AREAS_CRITICAS = 0.40   # Fracción del consumo total en áreas críticas (40 %)
DIAS_POR_MES = 30                # Días promedio por mes para conversiones mensuales
POTENCIA_PANEL_W = 400.0         # Potencia pico por panel (W)
HORAS_SOL_PICO = 5.2             # HSP promedio anual Barranquilla
W_POR_KW = 1000.0                # Factor de conversión vatios a kilovatios
AREA_POR_PANEL_M2 = 2.0          # Área aproximada por panel estándar de 400 W (m²)

# Ruta al archivo Excel con el consumo real de energía 2018
_RUTA_EXCEL = os.path.join(
    os.path.dirname(__file__), "..", "..", "data", "CONSUMO DE ENERGIA 2018.xlsx"
)

# Nombres de las hojas del Excel (Enero a Noviembre)
_HOJAS_MESES = {
    "Enero":      "ENERO 2018",
    "Febrero":    "FEBRERO 2018 ",
    "Marzo":      "MARZO 2018",
    "Abril":      "ABRIL 2018",
    "Mayo":       "MAYO 2018",
    "Junio":      "JUNIO 2018",
    "Julio":      "JULIO 2018",
    "Agosto":     "AGOSTO 2018",
    "Septiembre": "SEPTIEMBRE 2018",
    "Octubre":    "OCTUBRE 2018",
    "Noviembre":  "NOVIEMBRE 2018 ",
}

# Irradiación solar global horizontal (GHI) mensual para Barranquilla (kWh/m²/día)
# Fuente: IDEAM – Atlas de Irradiación Solar de Colombia (2014) / NASA POWER
# Barranquilla · Lat: 10.97°N · Lon: 74.80°W
# Valores correspondientes al período 2018 (media multianual ajustada)
# Los valores reflejan las dos temporadas de lluvia del Caribe colombiano:
#   - 1ª temporada: Mayo–Junio (reducción de irradiación)
#   - Veranillo de San Juan: Julio–Agosto (recuperación)
#   - 2ª temporada: Septiembre–Noviembre (mínimos anuales en Oct)

# Anchos de columna para tablas de resultados
_COL_MES = 12
_COL_IRRAD = 20
_COL_GEN = 22
_COL_COB = 10
_COL_ESC = 22
_COL_GEN_ESC = 12

IRRADIANCIA_MENSUAL = {
    "Enero":      5.82,   # Temporada seca — máxima irradiación
    "Febrero":    5.95,   # Temporada seca — pico anual
    "Marzo":      5.87,   # Temporada seca mayor
    "Abril":      5.63,   # Cielos parcialmente nublados
    "Mayo":       4.91,   # Inicio 1ª temporada de lluvias
    "Junio":      4.78,   # 1ª temporada de lluvias
    "Julio":      5.31,   # Veranillo de San Juan
    "Agosto":     5.44,   # Veranillo de San Juan — buena irradiación
    "Septiembre": 4.62,   # 2ª temporada de lluvias
    "Octubre":    4.38,   # Mínimo absoluto anual
    "Noviembre":  4.55,   # Final de lluvias — recuperación
    "Diciembre":  5.70,   # Retorno temporada seca
}


# ---------------------------------------------------------------------------
# Lectura de datos reales del Excel
# ---------------------------------------------------------------------------

def cargar_consumo_nazareth(ruta_excel: str) -> dict[str, float]:
    """Lee el Excel de consumo 2018 y retorna el consumo mensual total
    de Hospital Nazareth 1 (sede 1), de Enero a Noviembre.

    La función detecta dinámicamente la fila de encabezado y la columna
    'CONSUMO (kWh)' en cada hoja, ya que la estructura varía entre meses.

    Args:
        ruta_excel: Ruta al archivo CONSUMO DE ENERGIA 2018.xlsx.

    Returns:
        Diccionario {mes: consumo_total_kWh} para los 11 meses disponibles.
    """
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    consumo_mensual: dict[str, float] = {}

    for mes, nombre_hoja in _HOJAS_MESES.items():
        ws = wb[nombre_hoja]

        # Buscar fila de encabezado y columna de consumo dinámicamente
        consumo_col: int | None = None
        header_row: int | None = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if any(
                isinstance(c, str) and c.strip().upper() in ("SEDES", "SEDE")
                for c in row
            ):
                for col_idx, cell in enumerate(row):
                    if isinstance(cell, str) and "CONSUMO" in cell.upper():
                        consumo_col = col_idx
                        header_row = row_idx
                        break
                if header_row is not None:
                    break

        if header_row is None or consumo_col is None:
            consumo_mensual[mes] = 0.0
            continue

        # Sumar consumo de Hospital Nazareth 1 (sede 1 únicamente)
        # (header_row + 2 omite la fila de unidades '(kWh)' que sigue al encabezado)
        total = 0.0
        for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
            if not row or len(row) <= consumo_col:
                continue
            sede = row[0]
            if not isinstance(sede, str) or sede.strip().upper() != SEDE_OBJETIVO:
                continue
            consumo = row[consumo_col]
            if isinstance(consumo, (int, float)) and consumo > 0:
                total += consumo

        consumo_mensual[mes] = total

    return consumo_mensual


# ---------------------------------------------------------------------------
# Funciones del modelo
# ---------------------------------------------------------------------------

def calcular_energia_diaria(irradiancia: float, area: float, eficiencia: float) -> float:
    """Calcula la energía solar generada en un día (kWh/día).

    Args:
        irradiancia: Irradiancia solar promedio diaria (kWh/m²/día).
        area: Área de paneles solares disponible (m²).
        eficiencia: Eficiencia del sistema fotovoltaico (0‒1).

    Returns:
        Energía generada en kWh/día.
    """
    return irradiancia * area * eficiencia


def calcular_cobertura(energia_generada: float, consumo: float) -> float:
    """Calcula el porcentaje del consumo hospitalario cubierto por la generación solar.

    Args:
        energia_generada: Energía fotovoltaica generada (kWh/día).
        consumo: Consumo energético diario del hospital (kWh/día).

    Returns:
        Porcentaje de cobertura (0‒100).
    """
    return min((energia_generada / consumo) * 100, 100.0)


def calcular_generacion_mensual(
    irradiancia_mensual: dict[str, float],
    area: float,
    eficiencia: float,
) -> dict[str, float]:
    """Calcula la generación diaria estimada para cada mes.

    Args:
        irradiancia_mensual: Diccionario {mes: irradiancia (kWh/m²/día)}.
        area: Área de paneles solares (m²).
        eficiencia: Eficiencia del sistema fotovoltaico (0‒1).

    Returns:
        Diccionario {mes: energía generada (kWh/día)}.
    """
    return {
        mes: calcular_energia_diaria(irr, area, eficiencia)
        for mes, irr in irradiancia_mensual.items()
    }


# ---------------------------------------------------------------------------
# Utilidad de renderizado HTML → PNG
# ---------------------------------------------------------------------------

def _render_html(html: str, out_path: str, width: int = 1400) -> None:
    """Renderiza HTML+CSS a PNG de alta resolución usando Playwright/Chromium."""
    with sync_playwright() as pw:
        browser = pw.chromium.launch()
        page    = browser.new_page(
            viewport={"width": width, "height": 900},
            device_scale_factor=2,
        )
        page.set_content(html, wait_until="networkidle")
        height = page.evaluate("() => document.documentElement.scrollHeight")
        page.set_viewport_size({"width": width, "height": height})
        page.screenshot(path=out_path, full_page=True)
        browser.close()
    print(f"  Gráfico guardado: {out_path}")


_BASE_CSS = """
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
  background: #0f1729;
  color: #e2e8f0;
  -webkit-font-smoothing: antialiased;
}
"""

# ---------------------------------------------------------------------------
# Visualización — generacion_solar_nazareth.png
# ---------------------------------------------------------------------------

def graficar_generacion_mensual(
    generacion: dict[str, float],
    consumo: float,
    area: float,
    eficiencia: float,
) -> None:
    """Genera generacion_solar_nazareth.png con dos paneles:
    1. Generación solar diaria estimada por mes vs. consumo.
    2. Porcentaje de cobertura solar mensual.
    """
    meses    = list(generacion.keys())
    valores  = list(generacion.values())
    coberturas = [calcular_cobertura(v, consumo) for v in valores]
    max_val  = max(valores) * 1.2
    max_cob  = 115

    # ── barras gráfico 1 ──────────────────────────────────────────────────
    bars1 = ""
    for mes, val, cob in zip(meses, valores, coberturas):
        color = "#22c55e" if val >= consumo else "#3b82f6"
        h_pct = val / max_val * 100
        bars1 += f"""
        <div class="bar-col">
          <div class="bar-val">{val:,.0f}</div>
          <div class="bar-wrap">
            <div class="bar" style="height:{h_pct:.1f}%;background:{color}"></div>
          </div>
          <div class="bar-label">{mes[:3]}</div>
        </div>"""

    # línea de consumo: posición como % desde arriba = (1 - consumo/max_val)*100
    consumo_top = (1 - consumo / max_val) * 100

    # ── barras gráfico 2 ──────────────────────────────────────────────────
    bars2 = ""
    for mes, cob in zip(meses, coberturas):
        color = "#22c55e" if cob >= 100 else "#f59e0b" if cob >= 50 else "#ef4444"
        h_pct = cob / max_cob * 100
        bars2 += f"""
        <div class="bar-col">
          <div class="bar-val" style="color:{color}">{cob:.1f}%</div>
          <div class="bar-wrap">
            <div class="bar" style="height:{h_pct:.1f}%;background:{color}"></div>
          </div>
          <div class="bar-label">{mes[:3]}</div>
        </div>"""

    # línea 100% en gráfico 2
    cien_top = (1 - 100 / max_cob) * 100

    html = f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<style>
{_BASE_CSS}
.page {{ width:1400px; padding:0 0 32px; }}

.page-header {{
  background: #1a3a5c;
  padding: 22px 40px;
  border-bottom: 3px solid #0ea5e9;
}}
.page-header h1 {{ font-size:20px; font-weight:700; color:#fff; }}
.page-header p  {{ font-size:12px; color:#94a3b8; margin-top:4px; }}

.chart-block {{
  margin: 24px 40px 0;
  background: #1a2540;
  border-radius: 14px;
  padding: 24px 28px 20px;
}}
.chart-title {{
  font-size: 14px;
  font-weight: 700;
  color: #e2e8f0;
  margin-bottom: 18px;
}}
.chart-subtitle {{ font-size: 12px; color: #94a3b8; margin-top: 4px; }}

/* Layout de barras */
.bars-area {{
  position: relative;
  height: 240px;
}}
.bars-row {{
  display: flex;
  align-items: flex-end;
  gap: 8px;
  height: 200px;
  padding: 0 0 0 0;
}}
.bar-col {{
  flex: 1;
  display: flex;
  flex-direction: column;
  align-items: center;
  gap: 4px;
  height: 100%;
}}
.bar-val {{
  font-size: 11px;
  color: #94a3b8;
  font-weight: 600;
  min-height: 16px;
  text-align: center;
}}
.bar-wrap {{
  flex: 1;
  width: 100%;
  display: flex;
  align-items: flex-end;
}}
.bar {{
  width: 100%;
  border-radius: 4px 4px 0 0;
  min-height: 4px;
  transition: height .3s;
}}
.bar-label {{
  font-size: 11px;
  color: #64748b;
  text-align: center;
  padding-top: 6px;
}}

/* Línea de referencia */
.ref-line {{
  position: absolute;
  left: 0; right: 0;
  border-top: 2px dashed #ef4444;
  pointer-events: none;
}}
.ref-label {{
  position: absolute;
  right: 0;
  font-size: 10px;
  color: #ef4444;
  font-weight: 600;
  background: #1a2540;
  padding: 0 4px;
  transform: translateY(-50%);
}}

/* Leyenda */
.legend {{
  display: flex;
  gap: 20px;
  margin-top: 14px;
}}
.legend-item {{
  display: flex;
  align-items: center;
  gap: 7px;
  font-size: 12px;
  color: #94a3b8;
}}
.legend-dot {{
  width: 12px; height: 12px;
  border-radius: 3px;
  flex-shrink: 0;
}}
</style>
</head><body>
<div class="page">
  <div class="page-header">
    <h1>Hospital Nazareth 1 — Modelo de Generación Solar · Barranquilla 2018</h1>
    <p>Área instalada: {area:,.0f} m²  ·  Eficiencia del sistema: {eficiencia*100:.0f}%  ·  Irradiación: IDEAM / NASA POWER</p>
  </div>

  <!-- Gráfico 1 -->
  <div class="chart-block">
    <div class="chart-title">Generación Solar Diaria Estimada por Mes
      <span class="chart-subtitle"> — kWh/día</span>
    </div>
    <div class="bars-area">
      <div class="ref-line" style="top:{consumo_top:.1f}%">
        <span class="ref-label">Consumo diario {consumo:,.0f} kWh/día</span>
      </div>
      <div class="bars-row">{bars1}</div>
    </div>
    <div class="legend">
      <div class="legend-item"><div class="legend-dot" style="background:#22c55e"></div>Supera el consumo</div>
      <div class="legend-item"><div class="legend-dot" style="background:#3b82f6"></div>Por debajo del consumo</div>
      <div class="legend-item"><div class="legend-dot" style="background:#ef4444;height:2px;border-radius:0;width:20px"></div>Consumo diario de referencia</div>
    </div>
  </div>

  <!-- Gráfico 2 -->
  <div class="chart-block">
    <div class="chart-title">Porcentaje de Cobertura del Consumo Hospitalario
      <span class="chart-subtitle"> — %</span>
    </div>
    <div class="bars-area">
      <div class="ref-line" style="top:{cien_top:.1f}%">
        <span class="ref-label">100% cobertura</span>
      </div>
      <div class="bars-row">{bars2}</div>
    </div>
    <div class="legend">
      <div class="legend-item"><div class="legend-dot" style="background:#22c55e"></div>Cobertura total (≥ 100%)</div>
      <div class="legend-item"><div class="legend-dot" style="background:#f59e0b"></div>Cobertura parcial (50–99%)</div>
      <div class="legend-item"><div class="legend-dot" style="background:#ef4444"></div>Cobertura crítica (< 50%)</div>
    </div>
  </div>
</div>
</body></html>"""

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "generacion_solar_nazareth.png")
    _render_html(html, out)


# ---------------------------------------------------------------------------
# Visualización — escenarios_instalacion_nazareth.png
# ---------------------------------------------------------------------------

def analizar_escenarios(consumo: float) -> None:
    """Compara tres escenarios de instalación y genera escenarios_instalacion_nazareth.png."""
    irradiancia_promedio = sum(IRRADIANCIA_MENSUAL.values()) / len(IRRADIANCIA_MENSUAL)
    escenarios = [
        {"nombre": "Pequeño",  "area": 200,  "color": "#3b82f6", "bg": "#1e3a5f"},
        {"nombre": "Mediano",  "area": 500,  "color": "#22c55e", "bg": "#14432a"},
        {"nombre": "Grande",   "area": 1000, "color": "#a855f7", "bg": "#3b1f5e"},
    ]
    for e in escenarios:
        e["energia"] = calcular_energia_diaria(irradiancia_promedio, e["area"], EFICIENCIA_SISTEMA)
        e["cobertura"] = calcular_cobertura(e["energia"], consumo)

    max_e = max(e["energia"] for e in escenarios) * 1.2

    print("\n" + "=" * 55)
    print("  COMPARACIÓN DE ESCENARIOS – IRRADIANCIA PROMEDIO ANUAL")
    print("=" * 55)
    print(f"  Irradiancia promedio anual: {irradiancia_promedio:.2f} kWh/m²/día")
    print(f"  Eficiencia del sistema:     {EFICIENCIA_SISTEMA * 100:.0f} %")
    print(f"  Consumo diario hospital:    {consumo:,.0f} kWh/día")
    print("-" * 55)
    print(f"  {'Escenario':<{_COL_ESC}} {'Generación':>{_COL_GEN_ESC}}  {'Cobertura':>{_COL_COB}}")
    print("-" * 55)
    for e in escenarios:
        label = f"{e['nombre']} ({e['area']} m²)"
        print(f"  {label:<{_COL_ESC}} {e['energia']:>{_COL_GEN_ESC}.1f} kWh  {e['cobertura']:>{_COL_COB - 2}.1f} %")
    print("=" * 55)

    consumo_top = (1 - consumo / max_e) * 100

    cards_html = ""
    bars_html  = ""
    for e in escenarios:
        h_pct = e["energia"] / max_e * 100
        cob_color = "#22c55e" if e["cobertura"] >= 100 else "#f59e0b" if e["cobertura"] >= 50 else "#ef4444"
        cards_html += f"""
        <div class="esc-card" style="border-top:4px solid {e['color']}">
          <div class="esc-area" style="color:{e['color']}">{e['area']} m²</div>
          <div class="esc-nombre">{e['nombre']}</div>
          <div class="esc-energia">{e['energia']:,.0f} <span>kWh/día</span></div>
          <div class="esc-cob" style="color:{cob_color}">{e['cobertura']:.1f}% cobertura</div>
          <div class="esc-paneles">{int(e['area'] / 2)} paneles · {e['area']} m²</div>
        </div>"""
        bars_html += f"""
        <div class="bar-col">
          <div class="bar-val">{e['energia']:,.0f} kWh/día</div>
          <div class="bar-wrap">
            <div class="bar" style="height:{h_pct:.1f}%;background:{e['color']}"></div>
          </div>
          <div class="bar-label">{e['nombre']}<br><small>{e['area']} m²</small></div>
        </div>"""

    html = f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<style>
{_BASE_CSS}
.page {{ width:1400px; padding:0 0 40px; }}
.page-header {{
  background: #1a3a5c;
  padding: 22px 40px;
  border-bottom: 3px solid #a855f7;
}}
.page-header h1 {{ font-size:20px; font-weight:700; color:#fff; }}
.page-header p  {{ font-size:12px; color:#94a3b8; margin-top:4px; }}

.content {{ padding: 24px 40px; display: flex; flex-direction: column; gap: 24px; }}

/* Tarjetas resumen */
.cards-row {{
  display: grid;
  grid-template-columns: repeat(3, 1fr);
  gap: 20px;
}}
.esc-card {{
  background: #1a2540;
  border-radius: 12px;
  padding: 24px 22px;
}}
.esc-area   {{ font-size:30px; font-weight:800; }}
.esc-nombre {{ font-size:14px; color:#94a3b8; margin: 4px 0 14px; text-transform:uppercase; letter-spacing:.08em; }}
.esc-energia {{ font-size:26px; font-weight:700; color:#e2e8f0; }}
.esc-energia span {{ font-size:13px; color:#64748b; }}
.esc-cob    {{ font-size:14px; font-weight:700; margin: 8px 0 6px; }}
.esc-paneles {{ font-size:12px; color:#64748b; }}

/* Gráfico de barras */
.chart-block {{
  background: #1a2540;
  border-radius: 14px;
  padding: 24px 80px 20px;
}}
.chart-title {{ font-size:14px; font-weight:700; color:#e2e8f0; margin-bottom:18px; }}

.bars-area {{ position: relative; height: 280px; }}
.bars-row {{
  display: flex;
  align-items: flex-end;
  gap: 60px;
  height: 220px;
  justify-content: center;
}}
.bar-col {{
  width: 180px;
  display: flex;
  flex-direction: column;
  align-items: center;
  gap: 6px;
  height: 100%;
}}
.bar-val  {{ font-size:13px; color:#e2e8f0; font-weight:700; }}
.bar-wrap {{ flex:1; width:100%; display:flex; align-items:flex-end; }}
.bar      {{ width:100%; border-radius:6px 6px 0 0; min-height:4px; }}
.bar-label {{ font-size:12px; color:#64748b; text-align:center; padding-top:8px; line-height:1.5; }}
.bar-label small {{ font-size:11px; }}

.ref-line  {{ position:absolute; left:0; right:0; border-top:2px dashed #ef4444; pointer-events:none; }}
.ref-label {{ position:absolute; right:0; font-size:11px; color:#ef4444; font-weight:600;
              background:#1a2540; padding:0 6px; transform:translateY(-50%); }}

/* kpi inferior */
.kpi-row {{
  display: grid;
  grid-template-columns: repeat(3, 1fr);
  gap: 16px;
  background: #1a2540;
  border-radius: 12px;
  padding: 20px 24px;
}}
.kpi {{ text-align: center; }}
.kpi-val {{ font-size: 18px; font-weight: 700; color: #0ea5e9; }}
.kpi-lbl {{ font-size: 12px; color: #64748b; margin-top: 4px; }}
</style>
</head><body>
<div class="page">
  <div class="page-header">
    <h1>Hospital Nazareth 1 — Comparación de Escenarios de Instalación Solar</h1>
    <p>Irradiancia promedio anual Barranquilla: {irradiancia_promedio:.2f} kWh/m²/día  ·  Eficiencia: {EFICIENCIA_SISTEMA*100:.0f}%  ·  Consumo diario real: {consumo:,.0f} kWh/día</p>
  </div>
  <div class="content">
    <div class="cards-row">{cards_html}</div>
    <div class="chart-block">
      <div class="chart-title">Energía generada diaria por escenario vs. consumo real del hospital</div>
      <div class="bars-area">
        <div class="ref-line" style="top:{consumo_top:.1f}%">
          <span class="ref-label">Consumo diario {consumo:,.0f} kWh/día</span>
        </div>
        <div class="bars-row">{bars_html}</div>
      </div>
    </div>
    <div class="kpi-row">
      <div class="kpi">
        <div class="kpi-val">{irradiancia_promedio:.2f} kWh/m²/día</div>
        <div class="kpi-lbl">Irradiancia promedio anual · Barranquilla</div>
      </div>
      <div class="kpi">
        <div class="kpi-val">{consumo:,.0f} kWh/día</div>
        <div class="kpi-lbl">Consumo diario real · Hospital Nazareth 1</div>
      </div>
      <div class="kpi">
        <div class="kpi-val">{EFICIENCIA_SISTEMA*100:.0f}%</div>
        <div class="kpi-lbl">Eficiencia del sistema fotovoltaico</div>
      </div>
    </div>
  </div>
</div>
</body></html>"""

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "escenarios_instalacion_nazareth.png")
    _render_html(html, out)


# ---------------------------------------------------------------------------
# Punto de entrada principal
# ---------------------------------------------------------------------------

def main() -> None:
    # -----------------------------------------------------------------------
    # Cargar consumo real de Hospital Nazareth 1 (sede 1)
    # -----------------------------------------------------------------------
    consumo_nazareth = cargar_consumo_nazareth(_RUTA_EXCEL)

    if not consumo_nazareth:
        print("No se encontraron datos de consumo para el Hospital Nazareth.")
        return

    # Consumo diario real (promedio de los 11 meses disponibles, Enero-Noviembre)
    consumo_promedio_mensual = sum(consumo_nazareth.values()) / len(consumo_nazareth)
    consumo_diario_real = consumo_promedio_mensual / DIAS_POR_MES

    print("=" * 65)
    print("  CONSUMO REAL – HOSPITAL NAZARETH 1 (2018)")
    print("=" * 65)
    print(f"  {'Mes':<12} {'Consumo total (kWh)':>22} {'Áreas críticas (kWh)':>22}")
    print("-" * 65)
    for mes, consumo in consumo_nazareth.items():
        criticas = consumo * FRACCION_AREAS_CRITICAS
        print(f"  {mes:<12} {consumo:>22,.0f} {criticas:>22,.0f}")
    print("-" * 65)
    criticas_promedio = consumo_promedio_mensual * FRACCION_AREAS_CRITICAS
    print(f"  {'Promedio':<12} {consumo_promedio_mensual:>22,.0f} {criticas_promedio:>22,.0f}")
    print("=" * 65)

    # -----------------------------------------------------------------------
    # Modelo de generación solar usando consumo real
    # -----------------------------------------------------------------------
    print()
    print("=" * 55)
    print("  MODELO BÁSICO DE GENERACIÓN SOLAR – HOSPITAL NAZARETH")
    print("=" * 55)
    print(f"  Área de paneles:        {AREA_PANELES_M2:>8.0f} m²")
    print(f"  Eficiencia del sistema: {EFICIENCIA_SISTEMA * 100:>8.0f} %")
    print(f"  Consumo diario real:    {consumo_diario_real:>8,.0f} kWh/día")
    print(f"  (calculado de {len(consumo_nazareth)} meses, sede 1 únicamente)")
    print("-" * 55)
    print(f"  {'Mes':<{_COL_MES}} {'Irrad. (kWh/m²/día)':>{_COL_IRRAD}} {'Generación (kWh/día)':>{_COL_GEN}} {'Cobertura':>{_COL_COB}}")
    print("-" * 55)

    generacion = calcular_generacion_mensual(
        {m: IRRADIANCIA_MENSUAL[m] for m in consumo_nazareth},
        AREA_PANELES_M2,
        EFICIENCIA_SISTEMA,
    )

    for mes, energia in generacion.items():
        irr = IRRADIANCIA_MENSUAL[mes]
        consumo_diario_mes = consumo_nazareth[mes] / DIAS_POR_MES
        cobertura = calcular_cobertura(energia, consumo_diario_mes)
        print(f"  {mes:<{_COL_MES}} {irr:>{_COL_IRRAD}.1f} {energia:>{_COL_GEN}.1f} {cobertura:>{_COL_COB - 1}.1f} %")

    energia_promedio = sum(generacion.values()) / len(generacion)
    cobertura_promedio = calcular_cobertura(energia_promedio, consumo_diario_real)

    print("-" * 55)
    print(f"  {'Promedio':<{_COL_MES}} {'':>{_COL_IRRAD}} {energia_promedio:>{_COL_GEN}.1f} {cobertura_promedio:>{_COL_COB - 1}.1f} %")
    print("=" * 55)

    # -----------------------------------------------------------------------
    # Dimensionamiento del sistema para Sede 1
    # -----------------------------------------------------------------------
    energia_panel_kwh_dia = (POTENCIA_PANEL_W / W_POR_KW) * HORAS_SOL_PICO * EFICIENCIA_SISTEMA
    paneles_necesarios = math.ceil(consumo_diario_real / energia_panel_kwh_dia)
    area_necesaria_m2 = paneles_necesarios * AREA_POR_PANEL_M2

    print()
    print("=" * 55)
    print("  DIMENSIONAMIENTO DEL SISTEMA – HOSPITAL NAZARETH 1")
    print("=" * 55)
    print(f"  Consumo diario Sede 1:        {consumo_diario_real:>8,.0f} kWh/día")
    print(f"  Energía por panel/día:        {energia_panel_kwh_dia:>8.2f} kWh/panel/día")
    print(f"  Paneles necesarios (100 %):   {paneles_necesarios:>8} paneles")
    print(f"  Área mínima estimada:         {area_necesaria_m2:>8,.0f} m²")
    print("  (referencia para Actividad 03)")
    print("=" * 55)

    # -----------------------------------------------------------------------
    # Generación solar mensual propuesta (para comparar con consumo mensual)
    # -----------------------------------------------------------------------
    generacion_promedio_mensual = energia_promedio * DIAS_POR_MES

    print()
    print(
        f"El Hospital Nazareth 1 consume en promedio {consumo_promedio_mensual:,.0f} kWh "
        f"y el sistema solar propuesto generaría {generacion_promedio_mensual:,.0f} kWh"
    )
    print()

    analizar_escenarios(consumo_diario_real)
    graficar_generacion_mensual(
        generacion, consumo_diario_real, AREA_PANELES_M2, EFICIENCIA_SISTEMA
    )


if __name__ == "__main__":
    main()
