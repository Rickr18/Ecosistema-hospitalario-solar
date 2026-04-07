"""
modelos.py — Actividad 4: Dashboard energético Hospital Nazareth 1
Integra la lógica de generación solar (Actividad 2) y demanda
energética (Actividad 3). No duplica código: centraliza cálculos
para que el dashboard los consuma directamente.
"""

from __future__ import annotations

import calendar
import math
import os

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Constantes del sistema (heredadas de Actividades 2 y 3)
# ---------------------------------------------------------------------------
SEDE_OBJETIVO       = "HOSPITAL NAZARETH 1"
AREA_PANELES_M2     = 500.0       # m² instalados
EFICIENCIA_SISTEMA  = 0.15        # 15 %
DIAS_POR_MES        = 30
POTENCIA_PANEL_W    = 400.0
HORAS_SOL_PICO      = 5.2
W_POR_KW            = 1000.0
AREA_POR_PANEL_M2   = 2.0
TARIFA_COP_KWH      = 650.0       # COP/kWh (referencia gerencial)
FRACCION_CRITICAS   = 0.40        # 40 % del consumo en áreas críticas

# Irradiancia mensual Barranquilla (kWh/m²/día) — fuente IDEAM/NASA POWER
IRRADIANCIA_MENSUAL: dict[str, float] = {
    "Enero": 5.82, "Febrero": 5.95, "Marzo": 5.87,
    "Abril": 5.63, "Mayo": 4.91,   "Junio": 4.78,
    "Julio": 5.31, "Agosto": 5.44, "Septiembre": 4.62,
    "Octubre": 4.38, "Noviembre": 4.55,
}

MESES_ORDENADOS = list(IRRADIANCIA_MENSUAL.keys())

MESES_NUM_A_NOMBRE = {
    1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo",
    6:"Junio", 7:"Julio",  8:"Agosto", 9:"Septiembre",
    10:"Octubre", 11:"Noviembre",
}

# Hojas del Excel (nombres exactos del archivo institucional)
HOJAS_MESES: dict[str, str] = {
    "Enero":      "ENERO 2018",
    "Febrero":    "FEBRERO 2018 ",
    "Marzo":      "MARZO 2018",
    "Abril":      "ABRIL 2018",
    "Mayo":       "MAYO 2018",
    "Junio":      "JUNIO 2018",
    "Julio":      "JULIO 2018",
    "Agosto":     "AGOSTO 2018",
    "Septiembre": "SEPTIEMBRE 2018",
    "Octubre":    "OCTUBRE 2018",
    "Noviembre":  "NOVIEMBRE 2018 ",
}

# Pesos de consumo por área (Actividad 3 — deben sumar 1.0)
PESOS_AREAS: dict[str, float] = {
    "UCI":         0.19,
    "Quirófanos":  0.23,
    "Farmacia":    0.08,
    "Data Center": 0.11,
    "HVAC":        0.27,
    "Iluminación": 0.12,
}

COLORES_AREAS: dict[str, str] = {
    "UCI":         "#E24B4A",
    "Quirófanos":  "#f97316",
    "Farmacia":    "#EF9F27",
    "Data Center": "#7F77DD",
    "HVAC":        "#1D9E75",
    "Iluminación": "#378ADD",
}

# ---------------------------------------------------------------------------
# Lectura del Excel (lógica original Actividad 2)
# ---------------------------------------------------------------------------

def _ruta_excel_default() -> str:
    base = os.path.dirname(__file__)
    candidatos = [
        os.path.join(base, "CONSUMO DE ENERGIA 2018.xlsx"),
        os.path.join(base, "..", "data", "CONSUMO DE ENERGIA 2018.xlsx"),
        os.path.join(base, "..", "..", "data", "CONSUMO DE ENERGIA 2018.xlsx"),
    ]
    for ruta in candidatos:
        if os.path.exists(ruta):
            return os.path.normpath(ruta)
    raise FileNotFoundError(
        "No se encontró 'CONSUMO DE ENERGIA 2018.xlsx'. "
        "Colócalo en la carpeta data/ del proyecto o junto a este script."
    )


def cargar_consumo_mensual(ruta_excel: str | None = None) -> dict[str, float]:
    """
    Lee el Excel institucional 2018 y retorna el consumo mensual (kWh)
    de Hospital Nazareth 1, de Enero a Noviembre.
    Reutiliza la lógica de detección dinámica de columnas de la Actividad 2.
    """
    import openpyxl

    ruta = ruta_excel or _ruta_excel_default()
    wb = openpyxl.load_workbook(ruta, data_only=True)
    consumo: dict[str, float] = {}

    for mes, nombre_hoja in HOJAS_MESES.items():
        ws = wb[nombre_hoja]
        consumo_col: int | None = None
        header_row: int | None = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            if any(isinstance(c, str) and c.strip().upper() in ("SEDES", "SEDE") for c in row):
                for col_idx, cell in enumerate(row):
                    if isinstance(cell, str) and "CONSUMO" in cell.upper():
                        consumo_col = col_idx
                        header_row = row_idx
                        break
                if header_row is not None:
                    break

        if header_row is None or consumo_col is None:
            consumo[mes] = 0.0
            continue

        total = 0.0
        for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
            if not row or len(row) <= consumo_col:
                continue
            sede = row[0]
            valor = row[consumo_col]
            if (
                isinstance(sede, str)
                and sede.strip().upper() == SEDE_OBJETIVO
                and isinstance(valor, (int, float))
                and valor > 0
            ):
                total += float(valor)

        consumo[mes] = total

    return consumo


# ---------------------------------------------------------------------------
# Generación solar (Actividad 2)
# ---------------------------------------------------------------------------

def calcular_generacion_mensual(
    area: float = AREA_PANELES_M2,
    eficiencia: float = EFICIENCIA_SISTEMA,
    dias: int = DIAS_POR_MES,
) -> dict[str, float]:
    """Retorna la energía solar generada por mes (kWh/mes)."""
    return {
        mes: irr * area * eficiencia * dias
        for mes, irr in IRRADIANCIA_MENSUAL.items()
    }


def calcular_escenarios(consumo_diario: float) -> dict[str, dict]:
    """Compara tres escenarios de área instalada."""
    irr_prom = sum(IRRADIANCIA_MENSUAL.values()) / len(IRRADIANCIA_MENSUAL)
    escenarios = {
        "Pequeño (200 m²)":  200,
        "Mediano (500 m²)":  500,
        "Grande (1 000 m²)": 1000,
    }
    resultado = {}
    for nombre, area in escenarios.items():
        gen_dia = irr_prom * area * EFICIENCIA_SISTEMA
        cob = min((gen_dia / consumo_diario) * 100, 100) if consumo_diario else 0
        resultado[nombre] = {
            "area_m2": area,
            "generacion_diaria_kwh": round(gen_dia, 1),
            "cobertura_pct": round(cob, 1),
        }
    return resultado


def dimensionamiento(consumo_diario: float) -> dict:
    """Calcula paneles necesarios para cubrir el 100% del consumo."""
    energia_panel = (POTENCIA_PANEL_W / W_POR_KW) * HORAS_SOL_PICO * EFICIENCIA_SISTEMA
    paneles = math.ceil(consumo_diario / energia_panel)
    return {
        "consumo_diario_kwh": round(consumo_diario, 1),
        "energia_panel_kwh_dia": round(energia_panel, 3),
        "paneles_necesarios": paneles,
        "area_minima_m2": paneles * AREA_POR_PANEL_M2,
    }


# ---------------------------------------------------------------------------
# KPIs gerenciales (nuevos en Actividad 4)
# ---------------------------------------------------------------------------

def calcular_kpis_mes(mes: str, consumo_mensual: dict[str, float]) -> dict:
    """
    Calcula todos los KPIs gerenciales para un mes dado.
    Combina generación (Act. 2) y consumo real (Excel).
    """
    irr = IRRADIANCIA_MENSUAL[mes]
    gen_kwh = irr * AREA_PANELES_M2 * EFICIENCIA_SISTEMA * DIAS_POR_MES
    cons_kwh = consumo_mensual.get(mes, 0.0)

    autosuf = min((gen_kwh / cons_kwh) * 100, 100.0) if cons_kwh else 0.0
    excedente = max(0.0, gen_kwh - cons_kwh)
    deficit = max(0.0, cons_kwh - gen_kwh)
    ahorro_cop = gen_kwh * TARIFA_COP_KWH
    consumo_diario = cons_kwh / DIAS_POR_MES
    gen_diaria = gen_kwh / DIAS_POR_MES

    return {
        "mes":              mes,
        "generacion_kwh":   round(gen_kwh, 1),
        "consumo_kwh":      round(cons_kwh, 1),
        "autosuficiencia":  round(autosuf, 2),
        "excedente_kwh":    round(excedente, 1),
        "deficit_kwh":      round(deficit, 1),
        "ahorro_cop":       round(ahorro_cop, 0),
        "consumo_diario":   round(consumo_diario, 1),
        "gen_diaria":       round(gen_diaria, 1),
        "irradiancia":      irr,
    }


def calcular_kpis_anuales(consumo_mensual: dict[str, float]) -> pd.DataFrame:
    """Retorna DataFrame con KPIs de todos los meses disponibles."""
    filas = [calcular_kpis_mes(m, consumo_mensual) for m in MESES_ORDENADOS]
    return pd.DataFrame(filas).set_index("mes")


def calcular_areas_mes(consumo_mes_kwh: float) -> dict[str, float]:
    """Distribuye el consumo mensual por área según pesos de Actividad 3."""
    return {area: round(consumo_mes_kwh * peso, 1) for area, peso in PESOS_AREAS.items()}


def generar_alertas(kpis: dict) -> list[dict]:
    """
    Genera lista de alertas gerenciales basadas en umbrales.
    Cada alerta tiene: nivel ('ok'|'warn'|'critical'), mensaje.
    """
    alertas = []
    a = kpis["autosuficiencia"]
    d = kpis["deficit_kwh"]
    e = kpis["excedente_kwh"]

    if d > 0:
        alertas.append({
            "nivel": "critical",
            "mensaje": f"Déficit energético: {d:,.0f} kWh — Red eléctrica activa este mes.",
        })
    else:
        alertas.append({
            "nivel": "ok",
            "mensaje": f"Sin sobredemanda. Excedente solar: {e:,.0f} kWh.",
        })

    if a < 10:
        alertas.append({
            "nivel": "critical",
            "mensaje": f"Autosuficiencia {a:.1f}% — Por debajo del umbral mínimo (10%). Evaluar ampliación.",
        })
    elif a < 15:
        alertas.append({
            "nivel": "warn",
            "mensaje": f"Autosuficiencia {a:.1f}% — Por debajo del objetivo gerencial (15%). Revisar escenario 1 000 m².",
        })
    else:
        alertas.append({
            "nivel": "ok",
            "mensaje": f"Autosuficiencia {a:.1f}% — Cumple umbral gerencial ≥ 15%.",
        })

    uci_dia = kpis["consumo_kwh"] * PESOS_AREAS["UCI"] / DIAS_POR_MES
    if uci_dia > 300:
        alertas.append({
            "nivel": "warn",
            "mensaje": f"UCI consume ~{uci_dia:,.0f} kWh/día — Monitoreo prioritario recomendado.",
        })

    return alertas
