"""
Actividad 3: Simulacion de la demanda energetica hospitalaria
Hospital Nazareth 1 - Barranquilla, Colombia

Este script construye una simulacion horaria de la demanda energetica por
areas hospitalarias criticas y la calibra con el consumo mensual real 2018
(Enero a Noviembre) del archivo Excel institucional.

Entregables que cubre:
- Simulacion por areas: UCI, Quirofanos, Farmacia, Data Center, HVAC, Iluminacion
- Graficos de perfiles diario, semanal y mensual
- Analisis de picos de demanda global y por area
- Reporte HTML/CSS interactivo con Chart.js
"""

from __future__ import annotations

import calendar
import json
import os
import sys
from dataclasses import dataclass


def validar_ejecucion_en_venv() -> None:
    """Detiene la ejecucion si no se usa la .venv del proyecto."""
    venv_python = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..", ".venv", "Scripts", "python.exe")
    )

    if os.name == "nt" and os.path.exists(venv_python):
        exe_actual = os.path.normcase(os.path.normpath(sys.executable))
        exe_venv = os.path.normcase(venv_python)
        if exe_actual != exe_venv:
            print("ERROR: este script debe ejecutarse con la .venv del proyecto.")
            print("Ejecuta:")
            print("  ..\\..\\.venv\\Scripts\\python.exe simulacion_demanda_energetica.py")
            raise SystemExit(1)


validar_ejecucion_en_venv()

import numpy as np
import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# Configuracion general
# ---------------------------------------------------------------------------
SEDE_OBJETIVO = "HOSPITAL NAZARETH 1"
ANIO_ANALISIS = 2018
MESES_ANALISIS = list(range(1, 12))  # Enero -> Noviembre
DIAS_SEMANA_ORDEN = [
    "Lunes",
    "Martes",
    "Miercoles",
    "Jueves",
    "Viernes",
    "Sabado",
    "Domingo",
]

# Pesos de consumo mensual por area (deben sumar 1.0)
PESOS_AREAS = {
    "UCI": 0.19,
    "Quirofanos": 0.23,
    "Farmacia": 0.08,
    "Data Center": 0.11,
    "HVAC": 0.27,
    "Iluminacion": 0.12,
}

# Nombres de hojas de Excel 2018
HOJAS_MESES = {
    "Enero": "ENERO 2018",
    "Febrero": "FEBRERO 2018 ",
    "Marzo": "MARZO 2018",
    "Abril": "ABRIL 2018",
    "Mayo": "MAYO 2018",
    "Junio": "JUNIO 2018",
    "Julio": "JULIO 2018",
    "Agosto": "AGOSTO 2018",
    "Septiembre": "SEPTIEMBRE 2018",
    "Octubre": "OCTUBRE 2018",
    "Noviembre": "NOVIEMBRE 2018 ",
}

MESES_NUM_A_NOMBRE = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
}

COLORES_AREAS = {
    "UCI": "#ef4444",
    "Quirofanos": "#f97316",
    "Farmacia": "#f59e0b",
    "Data Center": "#6366f1",
    "HVAC": "#10b981",
    "Iluminacion": "#3b82f6",
    "Total": "#111827",
}

# Colores con transparencia para Chart.js (barras apiladas)
COLORES_AREAS_BG = {
    "UCI": "rgba(239,68,68,0.85)",
    "Quirofanos": "rgba(249,115,22,0.85)",
    "Farmacia": "rgba(245,158,11,0.85)",
    "Data Center": "rgba(99,102,241,0.85)",
    "HVAC": "rgba(16,185,129,0.85)",
    "Iluminacion": "rgba(59,130,246,0.85)",
}


@dataclass(frozen=True)
class PerfilArea:
    nombre: str

    def factor_horario(self, hora: int, es_fin_semana: bool) -> float:
        """Retorna el factor relativo de demanda para una hora dada."""
        if self.nombre == "UCI":
            return 0.95 + 0.12 * np.sin((hora - 5) * np.pi / 12) ** 2

        if self.nombre == "Quirofanos":
            base = 0.2
            if 7 <= hora <= 18:
                base += 1.25
            elif 19 <= hora <= 21:
                base += 0.35
            if es_fin_semana:
                base *= 0.55
            return base

        if self.nombre == "Farmacia":
            base = 0.55
            if 6 <= hora <= 20:
                base += 0.4
            if es_fin_semana:
                base *= 0.9
            return base

        if self.nombre == "Data Center":
            base = 0.82
            if 8 <= hora <= 19:
                base += 0.28
            return base

        if self.nombre == "HVAC":
            base = 0.35
            if 8 <= hora <= 17:
                base += 1.15
            elif 18 <= hora <= 22:
                base += 0.55
            if es_fin_semana:
                base *= 0.85
            return base

        if self.nombre == "Iluminacion":
            base = 0.25
            if hora <= 6 or hora >= 18:
                base += 1.05
            elif 7 <= hora <= 17:
                base += 0.35
            return base

        return 1.0


def resolver_ruta_excel() -> str:
    """Prioriza Excel local de la actividad y usa fallback al dataset global."""
    base = os.path.dirname(__file__)
    local = os.path.join(base, "CONSUMO DE ENERGIA 2018.xlsx")
    global_path = os.path.join(base, "..", "..", "data", "CONSUMO DE ENERGIA 2018.xlsx")

    if os.path.exists(local):
        return local
    if os.path.exists(global_path):
        return global_path

    raise FileNotFoundError(
        "No se encontro 'CONSUMO DE ENERGIA 2018.xlsx' ni en la carpeta de la actividad "
        "ni en ../../data/."
    )


def cargar_consumo_nazareth(ruta_excel: str) -> dict[str, float]:
    """Lee consumo mensual total de Hospital Nazareth 1 desde el Excel 2018."""
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    consumo_mensual: dict[str, float] = {}

    for mes, nombre_hoja in HOJAS_MESES.items():
        ws = wb[nombre_hoja]

        consumo_col: int | None = None
        header_row: int | None = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if any(isinstance(c, str) and c.strip().upper() in ("SEDES", "SEDE") for c in row):
                for col_idx, cell in enumerate(row):
                    if isinstance(cell, str) and "CONSUMO" in cell.upper():
                        consumo_col = col_idx
                        header_row = row_idx
                        break
                if header_row is not None:
                    break

        if header_row is None or consumo_col is None:
            consumo_mensual[mes] = 0.0
            continue

        total = 0.0
        for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
            if not row or len(row) <= consumo_col:
                continue
            sede = row[0]
            consumo = row[consumo_col]
            if (
                isinstance(sede, str)
                and sede.strip().upper() == SEDE_OBJETIVO
                and isinstance(consumo, (int, float))
                and consumo > 0
            ):
                total += float(consumo)

        consumo_mensual[mes] = total

    return consumo_mensual


def construir_base_mensual(
    anio: int,
    mes: int,
    consumo_mensual_kwh: float,
    rng: np.random.Generator,
) -> pd.DataFrame:
    """Construye serie horaria de un mes y la calibra al consumo real mensual."""
    dias_en_mes = calendar.monthrange(anio, mes)[1]
    fechas = pd.date_range(
        start=f"{anio}-{mes:02d}-01 00:00:00",
        end=f"{anio}-{mes:02d}-{dias_en_mes:02d} 23:00:00",
        freq="h",
    )

    df = pd.DataFrame(index=fechas)

    for area, peso in PESOS_AREAS.items():
        perfil = PerfilArea(area)
        objetivo_area_mes = consumo_mensual_kwh * peso

        valores_raw = []
        for ts in fechas:
            hora = ts.hour
            es_fin_semana = ts.weekday() >= 5

            factor_hora = perfil.factor_horario(hora, es_fin_semana)
            ruido = rng.normal(loc=1.0, scale=0.06)
            ruido = max(0.78, min(1.22, ruido))
            factor_dia = 1.0 + 0.05 * np.sin((ts.day / dias_en_mes) * 2 * np.pi)

            valores_raw.append(factor_hora * ruido * factor_dia)

        serie_raw = np.array(valores_raw, dtype=float)
        total_raw = float(serie_raw.sum())
        escala = (objetivo_area_mes / total_raw) if total_raw > 0 else 0.0
        df[area] = serie_raw * escala

    df["Total"] = df[list(PESOS_AREAS.keys())].sum(axis=1)

    return df


def simular_demanda_horaria(consumo_mensual: dict[str, float], semilla: int = 42) -> pd.DataFrame:
    """Genera la simulacion horaria Ene-Nov calibrada con consumo real mensual."""
    rng = np.random.default_rng(semilla)
    marcos = []

    for mes in MESES_ANALISIS:
        nombre_mes = MESES_NUM_A_NOMBRE[mes]
        consumo_mes = consumo_mensual.get(nombre_mes, 0.0)
        marcos.append(construir_base_mensual(ANIO_ANALISIS, mes, consumo_mes, rng))

    df = pd.concat(marcos).sort_index()
    return df


def enriquecer_columnas_tiempo(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega columnas de apoyo temporal para analitica y graficos."""
    out = df.copy()
    out["Mes"] = out.index.month.map(MESES_NUM_A_NOMBRE)
    out["Hora"] = out.index.hour
    out["DiaSemana"] = out.index.weekday
    out["DiaSemanaNombre"] = out["DiaSemana"].map(dict(enumerate(DIAS_SEMANA_ORDEN)))
    return out


def calcular_picos(df: pd.DataFrame) -> tuple[pd.Series, dict[str, pd.Series], pd.DataFrame]:
    """Calcula pico global, picos por area y top 10 horas de mayor demanda."""
    fila_pico_global = df["Total"].idxmax()
    pico_global = pd.Series(
        {
            "fecha_hora": fila_pico_global,
            "demanda_kwh": df.loc[fila_pico_global, "Total"],
        }
    )

    picos_area: dict[str, pd.Series] = {}
    for area in PESOS_AREAS:
        idx = df[area].idxmax()
        picos_area[area] = pd.Series({"fecha_hora": idx, "demanda_kwh": df.loc[idx, area]})

    top10 = df[["Total"]].sort_values("Total", ascending=False).head(10).copy()
    top10["FechaHora"] = top10.index
    top10 = top10[["FechaHora", "Total"]]

    return pico_global, picos_area, top10


def imprimir_resumen(df: pd.DataFrame, consumo_mensual: dict[str, float]) -> None:
    """Imprime reporte resumido de calibracion y demanda simulada."""
    print("=" * 86)
    print("  ACTIVIDAD 3 - SIMULACION DEMANDA ENERGETICA HOSPITALARIA")
    print("  Hospital Nazareth 1 | Barranquilla | Base real 2018 (Ene-Nov)")
    print("=" * 86)

    print("\nConsumo mensual real calibrado (kWh):")
    for mes in MESES_ANALISIS:
        nombre = MESES_NUM_A_NOMBRE[mes]
        print(f"  - {nombre:<11}: {consumo_mensual.get(nombre, 0.0):>12,.0f}")

    energia_total = df["Total"].sum()
    promedio_horario = df["Total"].mean()
    promedio_diario = df["Total"].resample("D").sum().mean()

    print("\nIndicadores globales simulados:")
    print(f"  - Energia total simulada Ene-Nov: {energia_total:,.0f} kWh")
    print(f"  - Demanda horaria promedio:       {promedio_horario:,.2f} kWh/h")
    print(f"  - Demanda diaria promedio:        {promedio_diario:,.0f} kWh/dia")

    print("\nParticipacion de areas en la demanda total:")
    for area in PESOS_AREAS:
        participacion = (df[area].sum() / energia_total) * 100 if energia_total else 0.0
        print(f"  - {area:<12}: {participacion:>6.2f} %")


def imprimir_reporte_picos(
    pico_global: pd.Series,
    picos_area: dict[str, pd.Series],
    top10: pd.DataFrame,
) -> None:
    """Imprime el reporte de picos de demanda."""
    print("\n" + "=" * 86)
    print("  ANALISIS DE PICOS DE DEMANDA")
    print("=" * 86)

    fecha_global = pd.Timestamp(pico_global["fecha_hora"]).strftime("%Y-%m-%d %H:%M")
    print(f"Pico global: {fecha_global} -> {float(pico_global['demanda_kwh']):,.2f} kWh/h")

    print("\nPico por area:")
    for area, fila in picos_area.items():
        fecha = pd.Timestamp(fila["fecha_hora"]).strftime("%Y-%m-%d %H:%M")
        print(f"  - {area:<12}: {fecha} -> {float(fila['demanda_kwh']):,.2f} kWh/h")

    print("\nTop 10 horas de mayor demanda total:")
    for i, row in enumerate(top10.itertuples(index=False), start=1):
        fecha = pd.Timestamp(row.FechaHora).strftime("%Y-%m-%d %H:%M")
        print(f"  {i:>2}. {fecha} -> {float(row.Total):,.2f} kWh/h")


# ---------------------------------------------------------------------------
# Generacion del reporte HTML
# ---------------------------------------------------------------------------

def _preparar_datos_html(
    df: pd.DataFrame,
    consumo_mensual: dict[str, float],
    pico_global: pd.Series,
    picos_area: dict[str, pd.Series],
    top10: pd.DataFrame,
) -> dict:
    """Extrae y serializa todos los datos necesarios para el HTML."""
    areas = list(PESOS_AREAS.keys())
    energia_total = float(df["Total"].sum())
    promedio_horario = float(df["Total"].mean())
    promedio_diario = float(df["Total"].resample("D").sum().mean())
    pico_valor = float(pico_global["demanda_kwh"])
    pico_fecha = pd.Timestamp(pico_global["fecha_hora"]).strftime("%d/%m/%Y %H:%M")

    # --- Perfil diario ---
    ph = df.groupby("Hora")[areas + ["Total"]].mean()
    perfil_diario = {
        "labels": list(range(24)),
        "total": [round(ph.loc[h, "Total"], 2) for h in range(24)],
        "areas": {a: [round(ph.loc[h, a], 2) for h in range(24)] for a in areas},
    }

    # --- Perfil semanal ---
    diario = (
        df.groupby([df.index.date, "DiaSemanaNombre"])[areas]
        .sum()
        .reset_index()
    )
    diario = diario.rename(columns={diario.columns[0]: "Fecha"})
    ps = (
        diario.groupby("DiaSemanaNombre")[areas]
        .mean()
        .reindex(DIAS_SEMANA_ORDEN)
    )
    perfil_semanal = {
        "labels": DIAS_SEMANA_ORDEN,
        "areas": {a: [round(ps.loc[d, a], 2) for d in DIAS_SEMANA_ORDEN] for a in areas},
    }

    # --- Perfil mensual ---
    orden_meses = [MESES_NUM_A_NOMBRE[m] for m in MESES_ANALISIS]
    pm = df.groupby("Mes")[areas].sum().reindex(orden_meses)
    perfil_mensual = {
        "labels": orden_meses,
        "areas": {a: [round(float(pm.loc[mes, a]), 0) for mes in orden_meses] for a in areas},
        "consumo_real": [round(consumo_mensual.get(mes, 0.0), 0) for mes in orden_meses],
    }

    # --- Heatmap (muestra el promedio horario por mes) ---
    pivot_hm = (
        df.assign(MesNum=df.index.month)
        .groupby(["MesNum", "Hora"])["Total"]
        .mean()
        .unstack(level="Hora")
    )
    heatmap_rows = []
    for m_num in MESES_ANALISIS:
        row_vals = [round(float(pivot_hm.loc[m_num, h]), 2) for h in range(24)]
        heatmap_rows.append({"mes": MESES_NUM_A_NOMBRE[m_num], "valores": row_vals})

    heatmap_min = min(v for r in heatmap_rows for v in r["valores"])
    heatmap_max = max(v for r in heatmap_rows for v in r["valores"])

    # --- Top 10 ---
    top10_lista = [
        {
            "rank": i + 1,
            "fecha": pd.Timestamp(row.FechaHora).strftime("%d/%m/%Y %H:%M"),
            "demanda": round(float(row.Total), 2),
        }
        for i, row in enumerate(top10.itertuples(index=False))
    ]

    # --- Picos por area ---
    picos_lista = [
        {
            "area": area,
            "fecha": pd.Timestamp(fila["fecha_hora"]).strftime("%d/%m/%Y %H:%M"),
            "demanda": round(float(fila["demanda_kwh"]), 2),
            "color": COLORES_AREAS[area],
        }
        for area, fila in picos_area.items()
    ]

    # --- Participacion por area ---
    participacion = {
        a: round((df[a].sum() / energia_total) * 100, 2) if energia_total else 0.0
        for a in areas
    }

    return {
        "energia_total": round(energia_total, 0),
        "promedio_horario": round(promedio_horario, 2),
        "promedio_diario": round(promedio_diario, 0),
        "pico_valor": round(pico_valor, 2),
        "pico_fecha": pico_fecha,
        "perfil_diario": perfil_diario,
        "perfil_semanal": perfil_semanal,
        "perfil_mensual": perfil_mensual,
        "heatmap_rows": heatmap_rows,
        "heatmap_min": round(heatmap_min, 2),
        "heatmap_max": round(heatmap_max, 2),
        "top10": top10_lista,
        "picos_area": picos_lista,
        "participacion": participacion,
        "areas": areas,
        "colores": COLORES_AREAS,
        "colores_bg": COLORES_AREAS_BG,
    }


def generar_html(datos: dict, output_dir: str) -> str:
    """Genera el reporte HTML con Chart.js embebido."""
    d = datos
    j = json.dumps(d, ensure_ascii=False)

    def fmt_num(n: float) -> str:
        return f"{n:,.0f}".replace(",", ".")

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Actividad 3 &mdash; Demanda Energetica Hospitalaria</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --bg: #f0f4f8;
    --card: #ffffff;
    --header: #1e3a5f;
    --accent: #2563eb;
    --text: #1e293b;
    --muted: #64748b;
    --border: #e2e8f0;
    --radius: 12px;
    --shadow: 0 2px 12px rgba(0,0,0,.08);
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', system-ui, sans-serif; background: var(--bg); color: var(--text); }}

  /* ---- Header ---- */
  .page-header {{
    background: linear-gradient(135deg, #1e3a5f 0%, #2563eb 100%);
    color: #fff;
    padding: 2.5rem 2rem 2rem;
    text-align: center;
  }}
  .page-header h1 {{ font-size: 1.75rem; font-weight: 700; margin-bottom: .4rem; }}
  .page-header p  {{ font-size: .95rem; opacity: .85; }}

  /* ---- Layout ---- */
  .container {{ max-width: 1280px; margin: 0 auto; padding: 2rem 1.5rem; }}

  /* ---- Section titles ---- */
  .section-title {{
    font-size: 1.15rem; font-weight: 700; color: var(--header);
    border-left: 4px solid var(--accent);
    padding-left: .75rem; margin: 2.5rem 0 1.25rem;
  }}

  /* ---- KPI cards ---- */
  .kpi-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 1.25rem;
  }}
  .kpi-card {{
    background: var(--card);
    border-radius: var(--radius);
    box-shadow: var(--shadow);
    padding: 1.4rem 1.5rem;
    display: flex; flex-direction: column; gap: .35rem;
    border-top: 4px solid var(--accent);
  }}
  .kpi-card.red   {{ border-top-color: #ef4444; }}
  .kpi-card.green {{ border-top-color: #10b981; }}
  .kpi-card.orange{{ border-top-color: #f97316; }}
  .kpi-label {{ font-size: .78rem; font-weight: 600; text-transform: uppercase; color: var(--muted); letter-spacing: .04em; }}
  .kpi-value {{ font-size: 1.8rem; font-weight: 800; color: var(--text); line-height: 1.1; }}
  .kpi-unit  {{ font-size: .82rem; color: var(--muted); }}

  /* ---- Chart cards ---- */
  .chart-card {{
    background: var(--card);
    border-radius: var(--radius);
    box-shadow: var(--shadow);
    padding: 1.5rem;
    margin-bottom: 1.5rem;
  }}
  .chart-card h3 {{
    font-size: 1rem; font-weight: 700; color: var(--header);
    margin-bottom: 1rem; padding-bottom: .6rem;
    border-bottom: 1px solid var(--border);
  }}
  .chart-wrap {{ position: relative; }}
  .chart-wrap canvas {{ max-height: 340px; }}

  /* ---- 2-col grid for charts ---- */
  .charts-2col {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 1.5rem;
  }}
  @media (max-width: 820px) {{
    .charts-2col {{ grid-template-columns: 1fr; }}
  }}

  /* ---- Heatmap CSS ---- */
  .heatmap-container {{ overflow-x: auto; }}
  .heatmap-table {{
    width: 100%; border-collapse: separate; border-spacing: 3px;
    font-size: .78rem;
  }}
  .heatmap-table th {{
    font-weight: 600; color: var(--muted);
    padding: 2px 4px; text-align: center; white-space: nowrap;
  }}
  .heatmap-table th.mes-label {{
    text-align: right; padding-right: 8px; font-size: .8rem; color: var(--text);
  }}
  .heatmap-cell {{
    width: 38px; height: 28px; border-radius: 4px;
    text-align: center; vertical-align: middle;
    font-size: .68rem; font-weight: 600; color: #fff;
    cursor: default;
  }}
  .hm-legend {{
    display: flex; align-items: center; gap: .5rem;
    font-size: .78rem; color: var(--muted); margin-top: .75rem;
  }}
  .hm-gradient {{
    height: 14px; width: 200px;
    background: linear-gradient(to right, #fef3c7, #f97316, #b91c1c);
    border-radius: 4px;
  }}

  /* ---- Tabla top 10 ---- */
  .data-table {{
    width: 100%; border-collapse: collapse; font-size: .9rem;
  }}
  .data-table thead th {{
    background: #f1f5f9; color: var(--muted);
    font-size: .78rem; font-weight: 700; text-transform: uppercase;
    padding: .6rem .9rem; text-align: left; border-bottom: 2px solid var(--border);
  }}
  .data-table tbody tr:nth-child(even) {{ background: #f8fafc; }}
  .data-table tbody td {{
    padding: .55rem .9rem; border-bottom: 1px solid var(--border);
  }}
  .badge {{
    display: inline-flex; align-items: center; justify-content: center;
    width: 24px; height: 24px; border-radius: 50%;
    background: var(--accent); color: #fff; font-size: .75rem; font-weight: 700;
  }}
  .badge.gold   {{ background: #f59e0b; }}
  .badge.silver {{ background: #9ca3af; }}
  .badge.bronze {{ background: #b45309; }}

  /* ---- Picos por area ---- */
  .picos-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
    gap: 1rem;
  }}
  .pico-card {{
    border-radius: var(--radius);
    border: 1px solid var(--border);
    padding: 1rem 1.2rem;
    background: var(--card);
    box-shadow: var(--shadow);
    border-left: 5px solid;
  }}
  .pico-card .pico-area {{ font-weight: 700; font-size: .95rem; margin-bottom: .25rem; }}
  .pico-card .pico-val  {{ font-size: 1.4rem; font-weight: 800; }}
  .pico-card .pico-fecha{{ font-size: .8rem; color: var(--muted); margin-top: .2rem; }}

  /* ---- Participacion barras ---- */
  .part-row {{
    display: flex; align-items: center; gap: .75rem;
    margin-bottom: .6rem; font-size: .88rem;
  }}
  .part-label {{ width: 100px; text-align: right; font-weight: 600; color: var(--muted); }}
  .part-bar-bg {{
    flex: 1; height: 18px; background: #f1f5f9; border-radius: 99px; overflow: hidden;
  }}
  .part-bar {{ height: 100%; border-radius: 99px; }}
  .part-pct {{ width: 45px; text-align: right; font-weight: 700; }}

  /* ---- Footer ---- */
  footer {{
    text-align: center; font-size: .8rem; color: var(--muted);
    padding: 2rem 1rem; border-top: 1px solid var(--border); margin-top: 1rem;
  }}
</style>
</head>
<body>

<div class="page-header">
  <h1>Simulacion de Demanda Energetica Hospitalaria</h1>
  <p>Hospital Nazareth 1 &mdash; Barranquilla, Colombia &mdash; Base real 2018 (Enero&ndash;Noviembre)</p>
</div>

<div class="container">

  <!-- KPIs -->
  <h2 class="section-title">Indicadores Globales</h2>
  <div class="kpi-grid">
    <div class="kpi-card">
      <span class="kpi-label">Energia total simulada</span>
      <span class="kpi-value">{fmt_num(d['energia_total'])}</span>
      <span class="kpi-unit">kWh &mdash; Ene&ndash;Nov 2018</span>
    </div>
    <div class="kpi-card green">
      <span class="kpi-label">Demanda horaria promedio</span>
      <span class="kpi-value">{fmt_num(d['promedio_horario'])}</span>
      <span class="kpi-unit">kWh/hora</span>
    </div>
    <div class="kpi-card orange">
      <span class="kpi-label">Demanda diaria promedio</span>
      <span class="kpi-value">{fmt_num(d['promedio_diario'])}</span>
      <span class="kpi-unit">kWh/dia</span>
    </div>
    <div class="kpi-card red">
      <span class="kpi-label">Pico maximo global</span>
      <span class="kpi-value">{fmt_num(d['pico_valor'])}</span>
      <span class="kpi-unit">kWh/h &mdash; {d['pico_fecha']}</span>
    </div>
  </div>

  <!-- Participacion por area -->
  <h2 class="section-title">Participacion por Area</h2>
  <div class="chart-card">
    <h3>Distribucion porcentual del consumo total</h3>
    <div id="part-bars"></div>
    <script>
    (function() {{
      const datos = {j};
      const cont = document.getElementById('part-bars');
      datos.areas.forEach(area => {{
        const pct = datos.participacion[area];
        const color = datos.colores[area];
        cont.innerHTML += `
          <div class="part-row">
            <span class="part-label">${{area}}</span>
            <div class="part-bar-bg">
              <div class="part-bar" style="width:${{pct}}%;background:${{color}}"></div>
            </div>
            <span class="part-pct">${{pct.toFixed(1)}}%</span>
          </div>`;
      }});
    }})();
    </script>
  </div>

  <!-- Perfil diario -->
  <h2 class="section-title">Perfil Horario Diario Promedio</h2>
  <div class="chart-card">
    <h3>Demanda promedio por hora del dia &mdash; todas las areas</h3>
    <div class="chart-wrap"><canvas id="chartDiario"></canvas></div>
  </div>

  <!-- Semanal + Mensual en 2 cols -->
  <h2 class="section-title">Perfiles Semanal y Mensual</h2>
  <div class="charts-2col">
    <div class="chart-card">
      <h3>Demanda promedio por dia de la semana (kWh/dia)</h3>
      <div class="chart-wrap"><canvas id="chartSemanal"></canvas></div>
    </div>
    <div class="chart-card">
      <h3>Demanda total por mes, Ene&ndash;Nov 2018 (kWh/mes)</h3>
      <div class="chart-wrap"><canvas id="chartMensual"></canvas></div>
    </div>
  </div>

  <!-- Heatmap -->
  <h2 class="section-title">Mapa de Calor &mdash; Demanda Horaria Media por Mes</h2>
  <div class="chart-card">
    <h3>Intensidad de demanda (kWh/h) por hora del dia y mes</h3>
    <div class="heatmap-container">
      <table class="heatmap-table" id="heatmapTable"></table>
    </div>
    <div class="hm-legend">
      <span>Bajo</span>
      <div class="hm-gradient"></div>
      <span>Alto</span>
    </div>
  </div>

  <!-- Picos por area -->
  <h2 class="section-title">Picos de Demanda por Area</h2>
  <div class="picos-grid" id="picosGrid"></div>

  <!-- Top 10 -->
  <h2 class="section-title">Top 10 Horas de Mayor Demanda</h2>
  <div class="chart-card" style="padding:0;overflow:hidden;">
    <table class="data-table" id="top10Table">
      <thead><tr>
        <th>#</th><th>Fecha y Hora</th><th>Demanda (kWh/h)</th>
      </tr></thead>
      <tbody></tbody>
    </table>
  </div>

</div><!-- .container -->

<footer>
  Actividad 3 &mdash; Simulacion de Demanda Energetica &mdash; Hospital Nazareth 1 &mdash; 2018
</footer>

<script>
// ---- Datos globales ----
const D = {j};

// ---- Perfil diario (lineas) ----
(function() {{
  const ctx = document.getElementById('chartDiario').getContext('2d');
  const datasets = D.areas.map(a => ({{
    label: a,
    data: D.perfil_diario.areas[a],
    borderColor: D.colores[a],
    backgroundColor: 'transparent',
    borderWidth: 2,
    pointRadius: 2,
    tension: 0.35,
  }}));
  datasets.push({{
    label: 'Total hospital',
    data: D.perfil_diario.total,
    borderColor: D.colores.Total,
    backgroundColor: 'transparent',
    borderWidth: 3,
    pointRadius: 3,
    tension: 0.35,
    borderDash: [6,3],
  }});
  new Chart(ctx, {{
    type: 'line',
    data: {{ labels: D.perfil_diario.labels.map(h => h + ':00'), datasets }},
    options: {{
      responsive: true,
      maintainAspectRatio: true,
      plugins: {{ legend: {{ position: 'bottom', labels: {{ boxWidth: 14, font: {{ size: 11 }} }} }} }},
      scales: {{
        x: {{ grid: {{ color: '#f0f4f8' }} }},
        y: {{ grid: {{ color: '#f0f4f8' }}, title: {{ display: true, text: 'kWh/h' }} }},
      }},
    }},
  }});
}})();

// ---- Perfil semanal (barras apiladas) ----
(function() {{
  const ctx = document.getElementById('chartSemanal').getContext('2d');
  const datasets = D.areas.map(a => ({{
    label: a,
    data: D.perfil_semanal.areas[a],
    backgroundColor: D.colores_bg[a],
    borderWidth: 0,
  }}));
  new Chart(ctx, {{
    type: 'bar',
    data: {{ labels: D.perfil_semanal.labels, datasets }},
    options: {{
      responsive: true,
      maintainAspectRatio: true,
      plugins: {{ legend: {{ position: 'bottom', labels: {{ boxWidth: 14, font: {{ size: 11 }} }} }} }},
      scales: {{
        x: {{ stacked: true, grid: {{ display: false }} }},
        y: {{ stacked: true, title: {{ display: true, text: 'kWh/dia' }} }},
      }},
    }},
  }});
}})();

// ---- Perfil mensual (barras apiladas) ----
(function() {{
  const ctx = document.getElementById('chartMensual').getContext('2d');
  const datasets = D.areas.map(a => ({{
    label: a,
    data: D.perfil_mensual.areas[a],
    backgroundColor: D.colores_bg[a],
    borderWidth: 0,
  }}));
  new Chart(ctx, {{
    type: 'bar',
    data: {{ labels: D.perfil_mensual.labels, datasets }},
    options: {{
      responsive: true,
      maintainAspectRatio: true,
      plugins: {{ legend: {{ position: 'bottom', labels: {{ boxWidth: 14, font: {{ size: 11 }} }} }} }},
      scales: {{
        x: {{ stacked: true, grid: {{ display: false }} }},
        y: {{ stacked: true, title: {{ display: true, text: 'kWh/mes' }} }},
      }},
    }},
  }});
}})();

// ---- Heatmap CSS ----
(function() {{
  const tbl = document.getElementById('heatmapTable');
  const min = D.heatmap_min;
  const max = D.heatmap_max;

  // Cabecera horas
  let header = '<tr><th class="mes-label">Mes</th>';
  for (let h = 0; h < 24; h++) header += `<th>${{h.toString().padStart(2,'0')}}</th>`;
  header += '</tr>';
  tbl.innerHTML = header;

  // Funcion de interpolacion YlOrRd simplificada
  function heatColor(val) {{
    const t = Math.max(0, Math.min(1, (val - min) / (max - min || 1)));
    // De amarillo #fef3c7 -> naranja #f97316 -> rojo oscuro #b91c1c
    const stops = [
      [254, 243, 199],
      [251, 191, 36],
      [249, 115, 22],
      [220, 38, 38],
      [185, 28, 28],
    ];
    const idx = t * (stops.length - 1);
    const lo = Math.floor(idx);
    const hi = Math.min(lo + 1, stops.length - 1);
    const f = idx - lo;
    const r = Math.round(stops[lo][0] + f * (stops[hi][0] - stops[lo][0]));
    const g = Math.round(stops[lo][1] + f * (stops[hi][1] - stops[lo][1]));
    const b = Math.round(stops[lo][2] + f * (stops[hi][2] - stops[lo][2]));
    const textColor = t > 0.45 ? '#fff' : '#374151';
    return {{ bg: `rgb(${{r}},${{g}},${{b}})`, text: textColor }};
  }}

  D.heatmap_rows.forEach(row => {{
    let tr = `<tr><th class="mes-label">${{row.mes}}</th>`;
    row.valores.forEach(v => {{
      const c = heatColor(v);
      tr += `<td class="heatmap-cell" style="background:${{c.bg}};color:${{c.text}}" title="${{v}} kWh/h">${{v.toFixed(0)}}</td>`;
    }});
    tr += '</tr>';
    tbl.innerHTML += tr;
  }});
}})();

// ---- Picos por area ----
(function() {{
  const cont = document.getElementById('picosGrid');
  D.picos_area.forEach(p => {{
    cont.innerHTML += `
      <div class="pico-card" style="border-left-color:${{p.color}}">
        <div class="pico-area" style="color:${{p.color}}">${{p.area}}</div>
        <div class="pico-val">${{p.demanda.toLocaleString('es-CO', {{minimumFractionDigits:2}})}} <small style="font-size:.7em;font-weight:400">kWh/h</small></div>
        <div class="pico-fecha">${{p.fecha}}</div>
      </div>`;
  }});
}})();

// ---- Top 10 ----
(function() {{
  const tbody = document.querySelector('#top10Table tbody');
  D.top10.forEach(row => {{
    const badgeClass = row.rank === 1 ? 'badge gold' : row.rank === 2 ? 'badge silver' : row.rank === 3 ? 'badge bronze' : 'badge';
    tbody.innerHTML += `
      <tr>
        <td><span class="${{badgeClass}}">${{row.rank}}</span></td>
        <td>${{row.fecha}}</td>
        <td><strong>${{row.demanda.toLocaleString('es-CO', {{minimumFractionDigits:2}})}}</strong></td>
      </tr>`;
  }});
}})();
</script>
</body>
</html>"""

    out = os.path.join(output_dir, "reporte_demanda.html")
    with open(out, "w", encoding="utf-8") as fh:
        fh.write(html)
    return out


def main() -> None:
    ruta_excel = resolver_ruta_excel()
    consumo_mensual = cargar_consumo_nazareth(ruta_excel)

    if not consumo_mensual:
        raise RuntimeError("No fue posible cargar consumo mensual del hospital.")

    df = simular_demanda_horaria(consumo_mensual, semilla=42)
    df = enriquecer_columnas_tiempo(df)

    imprimir_resumen(df, consumo_mensual)

    pico_global, picos_area, top10 = calcular_picos(df)
    imprimir_reporte_picos(pico_global, picos_area, top10)

    out_dir = os.path.dirname(__file__)
    datos = _preparar_datos_html(df, consumo_mensual, pico_global, picos_area, top10)
    ruta_html = generar_html(datos, out_dir)

    print(f"\nReporte HTML generado: {ruta_html}")


if __name__ == "__main__":
    main()
